import csv
from openpyxl import load_workbook

# Load the Excel workbook
wb = load_workbook('../data/Данные_для_курсовои_Классическое_МО.xlsx')
sheet = wb.active

# Write to CSV
with open('../data/kursovik_data.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)